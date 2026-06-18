#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Programme de veille documentaire automatisée.
Lit une liste de sources (sources.csv), récupère les nouveautés via RSS/Atom
(ou scraping léger), journalise tout dans un fichier Excel et télécharge
optionnellement les PDF associés.

Usage :
    python veille.py                # exécution normale
    python veille.py --reset        # réinitialise l'historique (re-signale tout)

Dépendances : feedparser, openpyxl, requests
    pip install feedparser openpyxl requests
"""

import csv
import os
import re
import sys
import json
import time
import hashlib
import datetime as dt
from urllib.parse import urljoin, urlparse

import requests
import feedparser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configuration — adapte ces chemins à ton environnement
# --------------------------------------------------------------------------
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
SOURCES_CSV = os.path.join(BASE_DIR, "sources.csv")
JOURNAL_XLSX= os.path.join(BASE_DIR, "veille.xlsx")
JOURNAL_HTML= os.path.join(BASE_DIR, "index.html")        # page web de veille
JOURNAL_RSS = os.path.join(BASE_DIR, "veille.xml")        # flux RSS agrégé
# URL publique où sera hébergé le flux (GitHub Pages, ENT...). Sert à remplir
# les liens <atom:link> et <guid>. Adapte-la, ou laisse vide si usage local.
SITE_URL    = "https://loicschreiber.github.io/veille"
DONNEES_JSON= os.path.join(BASE_DIR, "donnees.json")      # historique complet (pour le HTML)
VUS_JSON    = os.path.join(BASE_DIR, "vus.json")           # mémoire anti-doublons
PDF_DIR     = os.path.join(BASE_DIR, "documents")          # PDF téléchargés
PAUSE_SEC   = 2                                            # politesse entre requêtes
TIMEOUT     = 30
HEADERS     = {"User-Agent": "Veille-Perso/1.0 (usage pedagogique)"}
MAX_ENTREES_PAR_SOURCE = 50                                # garde-fou

# --------------------------------------------------------------------------
def charger_vus():
    if os.path.exists(VUS_JSON):
        with open(VUS_JSON, encoding="utf-8") as f:
            return set(json.load(f))
    return set()

def sauver_vus(vus):
    with open(VUS_JSON, "w", encoding="utf-8") as f:
        json.dump(sorted(vus), f, ensure_ascii=False, indent=0)

def id_entree(lien, titre):
    """Identifiant stable d'une entrée, basé sur le lien (ou le titre à défaut)."""
    base = (lien or titre or "").strip()
    return hashlib.sha1(base.encode("utf-8")).hexdigest()

def nettoyer_html(texte):
    if not texte:
        return ""
    texte = re.sub(r"<[^>]+>", " ", texte)
    texte = re.sub(r"\s+", " ", texte)
    return texte.strip()[:500]

def nom_fichier_sur(titre, ext=".pdf"):
    base = re.sub(r"[^\w\s-]", "", titre, flags=re.UNICODE).strip()
    base = re.sub(r"\s+", "_", base)[:80]
    return (base or "document") + ext

def date_entree(e):
    for champ in ("published_parsed", "updated_parsed"):
        val = e.get(champ)
        if val:
            return dt.datetime(*val[:6]).strftime("%Y-%m-%d %H:%M")
    return ""

# --------------------------------------------------------------------------
def lire_sources():
    sources = []
    with open(SOURCES_CSV, encoding="utf-8") as f:
        for ligne in csv.DictReader(f, delimiter=";"):
            if ligne.get("nom", "").strip() and not ligne["nom"].startswith("#"):
                sources.append(ligne)
    return sources

def urls_candidates(url):
    """Génère des variantes d'URL de flux plausibles à partir d'une URL de base.
    Permet de résister aux petites différences de chemin entre installations
    (SPIP, WordPress, OpenEdition...)."""
    candidates = [url]
    u = url.rstrip("/")
    base = f"{urlparse(u).scheme}://{urlparse(u).netloc}"
    # Si on nous donne une page d'accueil, on tente les motifs classiques
    motifs = [
        "/spip.php?page=backend",   # SPIP
        "/feed/", "/feed", "/rss", "/rss.xml",  # WordPress / divers
        "/backend?format=rssarticles",          # OpenEdition
        "/articles.atom",                        # The Conversation
    ]
    for m in motifs:
        candidates.append(u + m)
        candidates.append(base + m)
    # Dédoublonnage en gardant l'ordre
    vu, sortie = set(), []
    for c in candidates:
        if c not in vu:
            vu.add(c); sortie.append(c)
    return sortie

def recuperer_rss(url, auto=True):
    """Retourne (entrées, statut). Si auto=True et que l'URL fournie ne donne
    rien, essaie automatiquement des variantes courantes."""
    a_tester = urls_candidates(url) if auto else [url]
    derniere_info = "aucune entrée"
    for cand in a_tester:
        flux = feedparser.parse(cand, request_headers=HEADERS)
        if flux.entries:
            entrees = []
            for e in flux.entries[:MAX_ENTREES_PAR_SOURCE]:
                entrees.append({
                    "titre":  nettoyer_html(e.get("title", "(sans titre)")),
                    "lien":   e.get("link", ""),
                    "date":   date_entree(e),
                    "resume": nettoyer_html(e.get("summary", "")),
                })
            note = "" if cand == url else f" [URL retenue: {cand}]"
            return entrees, f"RSS ok, {len(entrees)} entrées{note}"
        derniere_info = f"http {flux.get('status','?')}, bozo={getattr(flux,'bozo',0)}"
    return [], f"RSS vide ({derniere_info})"

def recuperer_scrape(url):
    """Scraping minimal : récupère les liens d'une page (best-effort, sans BS4)."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
    except Exception as ex:
        return [], str(ex)
    liens = re.findall(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', r.text, re.IGNORECASE | re.DOTALL)
    entrees = []
    vus_locaux = set()
    for href, libelle in liens:
        libelle = nettoyer_html(libelle)
        href_abs = urljoin(url, href)
        # On ne garde que des liens "articles" plausibles, non vus, avec un libellé
        if (len(libelle) > 25 and href_abs not in vus_locaux
                and urlparse(href_abs).netloc == urlparse(url).netloc):
            vus_locaux.add(href_abs)
            entrees.append({"titre": libelle, "lien": href_abs, "date": "", "resume": ""})
        if len(entrees) >= 20:
            break
    return entrees, None

def telecharger_pdf(lien, titre):
    """Tente de télécharger un PDF. Retourne le chemin local ou ''."""
    os.makedirs(PDF_DIR, exist_ok=True)
    cible = lien
    try:
        # Si le lien n'est pas directement un PDF, on tente une requête HEAD
        if not lien.lower().endswith(".pdf"):
            h = requests.head(lien, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
            if "application/pdf" not in h.headers.get("Content-Type", ""):
                return ""  # pas un PDF, on n'insiste pas
        r = requests.get(cible, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        if not r.content.startswith(b"%PDF"):
            return ""
        chemin = os.path.join(PDF_DIR, nom_fichier_sur(titre))
        with open(chemin, "wb") as f:
            f.write(r.content)
        return chemin
    except Exception:
        return ""

# --------------------------------------------------------------------------
def init_journal():
    """Crée le classeur Excel avec en-têtes si absent, sinon le charge."""
    if os.path.exists(JOURNAL_XLSX):
        return load_workbook(JOURNAL_XLSX)
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    entetes = ["Date de détection", "Source", "Date article",
               "Titre", "Lien", "Résumé", "PDF local"]
    ws.append(entetes)
    entete_fill = PatternFill("solid", start_color="1F3864")
    for col, _ in enumerate(entetes, 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = entete_fill
        c.alignment = Alignment(vertical="center")
    largeurs = [18, 28, 16, 55, 50, 60, 30]
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb

def ajouter_lignes(wb, nouvelles):
    ws = wb["Journal"]
    for n in nouvelles:
        ligne = [n["detecte"], n["source"], n["date"], n["titre"],
                 n["lien"], n["resume"], n["pdf"]]
        ws.append(ligne)
        r = ws.max_row
        ws.cell(row=r, column=4).font = Font(name="Arial")
        if n["lien"]:
            cell = ws.cell(row=r, column=5)
            cell.hyperlink = n["lien"]
            cell.font = Font(color="0563C1", underline="single", name="Arial")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

# --------------------------------------------------------------------------
# Historique complet (sert à régénérer la page HTML à chaque exécution)
# --------------------------------------------------------------------------
def charger_donnees():
    if os.path.exists(DONNEES_JSON):
        with open(DONNEES_JSON, encoding="utf-8") as f:
            return json.load(f)
    return []

def sauver_donnees(donnees):
    # On garde les 1000 entrées les plus récentes pour ne pas alourdir la page
    donnees = donnees[-1000:]
    with open(DONNEES_JSON, "w", encoding="utf-8") as f:
        json.dump(donnees, f, ensure_ascii=False, indent=0)

def generer_html(donnees, maj):
    """Génère une page web autonome (un seul fichier) affichant toute la veille.
    Les données sont injectées en JSON ; la recherche et les filtres tournent
    côté navigateur, sans aucune dépendance externe."""
    # Tri du plus récent au plus ancien
    donnees_triees = sorted(donnees, key=lambda d: d.get("detecte", ""), reverse=True)
    sources = sorted({d["source"] for d in donnees_triees})
    data_json = json.dumps(donnees_triees, ensure_ascii=False)

    html = HTML_GABARIT.replace("/*__DATA__*/", data_json)
    html = html.replace("__MAJ__", maj)
    html = html.replace("__NB__", str(len(donnees_triees)))
    html = html.replace("__NB_SOURCES__", str(len(sources)))
    with open(JOURNAL_HTML, "w", encoding="utf-8") as f:
        f.write(html)

def _xml_echappe(t):
    """Échappement XML minimal pour le contenu textuel."""
    if not t:
        return ""
    return (t.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))

def _date_rfc822(detecte, date_art):
    """Convertit nos dates ('YYYY-MM-DD HH:MM') en format RFC-822 attendu par RSS."""
    for source in (date_art, detecte):
        if source:
            try:
                d = dt.datetime.strptime(source[:16], "%Y-%m-%d %H:%M")
                return d.strftime("%a, %d %b %Y %H:%M:%S +0000")
            except ValueError:
                continue
    return dt.datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")

def generer_rss(donnees, maj):
    """Génère un flux RSS 2.0 unique agrégeant toutes les sources.
    À ajouter dans n'importe quel lecteur de flux (Thunderbird, Inoreader...)."""
    donnees_triees = sorted(donnees, key=lambda d: d.get("detecte", ""), reverse=True)[:200]
    lien_flux = f"{SITE_URL}/veille.xml" if SITE_URL else "veille.xml"
    lien_site = SITE_URL or ""
    maj_rfc = _date_rfc822(maj, "")

    items = []
    for d in donnees_triees:
        # guid stable basé sur le lien (ou titre), garantit le suivi lu/non-lu
        guid = hashlib.sha1((d.get("lien") or d.get("titre","")).encode("utf-8")).hexdigest()
        titre = _xml_echappe(d.get("titre", "(sans titre)"))
        lien  = _xml_echappe(d.get("lien", ""))
        # On préfixe la description par la source pour la repérer dans le lecteur
        desc  = _xml_echappe(f"[{d.get('source','')}] {d.get('resume','')}".strip())
        items.append(f"""    <item>
      <title>{titre}</title>
      <link>{lien}</link>
      <description>{desc}</description>
      <category>{_xml_echappe(d.get('source',''))}</category>
      <guid isPermaLink="false">{guid}</guid>
      <pubDate>{_date_rfc822(d.get('detecte',''), d.get('date',''))}</pubDate>
    </item>""")

    rss = f"""<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0" xmlns:atom="http://www.w3.org/2005/Atom">
  <channel>
    <title>Veille — Histoire, Géographie, Éducation</title>
    <link>{_xml_echappe(lien_site)}</link>
    <atom:link href="{_xml_echappe(lien_flux)}" rel="self" type="application/rss+xml"/>
    <description>Flux agrégé de {len({d['source'] for d in donnees_triees})} sources de veille.</description>
    <language>fr</language>
    <lastBuildDate>{maj_rfc}</lastBuildDate>
{chr(10).join(items)}
  </channel>
</rss>
"""
    with open(JOURNAL_RSS, "w", encoding="utf-8") as f:
        f.write(rss)

# --------------------------------------------------------------------------
def main():
    reset = "--reset" in sys.argv
    vus = set() if reset else charger_vus()
    sources = lire_sources()
    maintenant = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    nouvelles = []
    print(f"== Veille du {maintenant} — {len(sources)} sources ==\n")

    for s in sources:
        nom, url, typ = s["nom"], s["url_flux"], s.get("type", "rss").strip().lower()
        dl_pdf = s.get("telecharger_pdf", "non").strip().lower() in ("oui", "yes", "true", "1")
        try:
            if typ == "scrape":
                entrees, err = recuperer_scrape(url)
                statut = "scrape" + (f" / {err}" if err else f", {len(entrees)} liens")
            else:
                entrees, statut = recuperer_rss(url)
        except Exception as ex:
            print(f"  [ERREUR] {nom} : {ex}")
            time.sleep(PAUSE_SEC)
            continue

        n_neuf = 0
        for e in entrees:
            uid = id_entree(e["lien"], e["titre"])
            if uid in vus:
                continue
            vus.add(uid)
            n_neuf += 1
            pdf_local = ""
            if dl_pdf and e["lien"]:
                pdf_local = telecharger_pdf(e["lien"], e["titre"])
            nouvelles.append({
                "detecte": maintenant, "source": nom, "date": e["date"],
                "titre": e["titre"], "lien": e["lien"],
                "resume": e["resume"], "pdf": pdf_local,
            })
        print(f"  {nom:<35} {n_neuf:>3} nouveau(x)   [{statut}]")
        time.sleep(PAUSE_SEC)

    # Historique complet pour la page web
    donnees = [] if reset else charger_donnees()
    donnees.extend(nouvelles)

    if nouvelles:
        wb = init_journal()
        ajouter_lignes(wb, nouvelles)
        wb.save(JOURNAL_XLSX)
        sauver_vus(vus)
        sauver_donnees(donnees)
        generer_html(donnees, maintenant)
        generer_rss(donnees, maintenant)
        print(f"\n>> {len(nouvelles)} entrée(s) ajoutée(s)")
        print(f"   - Excel : {os.path.basename(JOURNAL_XLSX)}")
        print(f"   - Page web : {os.path.basename(JOURNAL_HTML)}")
        print(f"   - Flux RSS : {os.path.basename(JOURNAL_RSS)}")
    else:
        sauver_vus(vus)
        # On régénère quand même la page (met à jour la date, garde l'historique)
        generer_html(donnees, maintenant)
        generer_rss(donnees, maintenant)
        print("\n>> Aucune nouveauté. Page web et flux RSS mis à jour.")

# --------------------------------------------------------------------------
# Gabarit de la page web (un seul fichier, aucune dépendance externe).
# Les marqueurs /*__DATA__*/, __MAJ__, __NB__, __NB_SOURCES__ sont remplacés
# à la génération.
# --------------------------------------------------------------------------
HTML_GABARIT = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Veille documentaire</title>
<style>
  :root{
    --encre:#1a2238; --papier:#f7f5ef; --carte:#fffdf8;
    --trait:#e0d9c8; --sourd:#6b6557; --accent:#9c3326; --accent-doux:#f0e4df;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  body{
    font-family:"Iowan Old Style","Palatino Linotype",Georgia,serif;
    background:var(--papier); color:var(--encre);
    line-height:1.5; -webkit-font-smoothing:antialiased;
  }
  .bandeau{
    border-bottom:2px solid var(--encre);
    padding:2.4rem 1.5rem 1.6rem; max-width:920px; margin:0 auto;
  }
  .surtitre{
    font:600 .72rem/1 ui-sans-serif,system-ui,sans-serif;
    letter-spacing:.22em; text-transform:uppercase; color:var(--accent);
    margin-bottom:.7rem;
  }
  h1{font-size:2.5rem; font-weight:700; letter-spacing:-.01em; line-height:1.05}
  .meta{
    font:.8rem/1.4 ui-sans-serif,system-ui,sans-serif; color:var(--sourd);
    margin-top:.6rem;
  }
  .outils{
    max-width:920px; margin:1.4rem auto 0; padding:0 1.5rem;
    display:flex; flex-wrap:wrap; gap:.5rem; align-items:center;
  }
  #recherche{
    flex:1 1 240px; min-width:0; padding:.6rem .8rem;
    font:1rem/1 ui-sans-serif,system-ui,sans-serif;
    border:1px solid var(--trait); border-radius:2px; background:var(--carte);
    color:var(--encre);
  }
  #recherche:focus{outline:2px solid var(--accent); outline-offset:1px}
  .filtres{
    max-width:920px; margin:.8rem auto 0; padding:0 1.5rem;
    display:flex; flex-wrap:wrap; gap:.4rem;
  }
  .puce{
    font:.74rem/1 ui-sans-serif,system-ui,sans-serif; cursor:pointer;
    padding:.4rem .7rem; border:1px solid var(--trait); border-radius:100px;
    background:var(--carte); color:var(--sourd); transition:.15s;
    user-select:none;
  }
  .puce:hover{border-color:var(--accent)}
  .puce.active{background:var(--encre); color:var(--papier); border-color:var(--encre)}
  .flux{max-width:920px; margin:1.6rem auto 4rem; padding:0 1.5rem}
  .fiche{
    border-bottom:1px solid var(--trait); padding:1.3rem 0;
    display:grid; grid-template-columns:8.5rem 1fr; gap:.35rem 1.4rem;
  }
  .fiche .src{
    grid-column:1; grid-row:1/4; align-self:start;
    font:600 .68rem/1.3 ui-sans-serif,system-ui,sans-serif;
    text-transform:uppercase; letter-spacing:.04em;
    color:var(--accent); padding-top:.25rem;
  }
  .fiche h2{grid-column:2; font-size:1.16rem; font-weight:600; line-height:1.3}
  .fiche h2 a{color:var(--encre); text-decoration:none}
  .fiche h2 a:hover{text-decoration:underline; text-decoration-color:var(--accent)}
  .fiche .res{
    grid-column:2;
    font:.9rem/1.45 ui-sans-serif,system-ui,sans-serif;
    color:var(--sourd); margin-top:.1rem;
  }
  .fiche .bas{
    grid-column:2;
    font:.72rem/1 ui-sans-serif,system-ui,sans-serif; color:var(--sourd);
    margin-top:.15rem; display:flex; gap:1rem; flex-wrap:wrap;
  }
  .fiche .bas .pdf{color:var(--accent); font-weight:600; text-decoration:none}
  .vide{
    text-align:center; color:var(--sourd); padding:3rem 1rem;
    font:1rem/1.5 ui-sans-serif,system-ui,sans-serif;
  }
  @media (max-width:560px){
    h1{font-size:1.9rem}
    .fiche{grid-template-columns:1fr; gap:.3rem}
    .fiche .src{grid-row:auto; width:auto; padding-top:0}
  }
  @media (prefers-reduced-motion:reduce){*{transition:none!important}}
</style>
</head>
<body>
  <header class="bandeau">
    <div class="surtitre">Veille &mdash; Histoire, Géographie, Éducation</div>
    <h1>Le fil des publications</h1>
    <p class="meta">__NB__ entrées &middot; __NB_SOURCES__ sources &middot; dernière mise à jour le __MAJ__</p>
  </header>
  <div class="outils">
    <input id="recherche" type="search" placeholder="Rechercher un titre, un mot-clé…" aria-label="Rechercher">
  </div>
  <div class="filtres" id="filtres"></div>
  <main class="flux" id="flux"></main>

<script>
const DONNEES = /*__DATA__*/;
let filtreSource = null;

const flux = document.getElementById("flux");
const filtres = document.getElementById("filtres");
const recherche = document.getElementById("recherche");

// Construction des puces de filtre par source
const sources = [...new Set(DONNEES.map(d => d.source))].sort();
function puce(label, val){
  const b = document.createElement("button");
  b.className = "puce" + (val === filtreSource ? " active" : "");
  b.textContent = label;
  b.onclick = () => { filtreSource = (filtreSource === val ? null : val); rendre(); };
  return b;
}
function construireFiltres(){
  filtres.innerHTML = "";
  filtres.appendChild(puce("Toutes les sources", null));
  sources.forEach(s => filtres.appendChild(puce(s, s)));
}

function echapper(t){
  const d = document.createElement("div"); d.textContent = t || ""; return d.innerHTML;
}

function rendre(){
  construireFiltres();
  const q = recherche.value.trim().toLowerCase();
  const liste = DONNEES.filter(d => {
    if (filtreSource && d.source !== filtreSource) return false;
    if (q){
      const foin = (d.titre + " " + d.resume + " " + d.source).toLowerCase();
      return foin.includes(q);
    }
    return true;
  });
  if (liste.length === 0){
    flux.innerHTML = '<p class="vide">Aucune entrée ne correspond. Modifiez votre recherche ou réinitialisez les filtres.</p>';
    return;
  }
  flux.innerHTML = liste.map(d => {
    const titre = d.lien
      ? `<a href="${echapper(d.lien)}" target="_blank" rel="noopener">${echapper(d.titre)}</a>`
      : echapper(d.titre);
    const res = d.resume ? `<p class="res">${echapper(d.resume)}</p>` : "";
    const dateArt = d.date ? echapper(d.date) : "détecté le " + echapper(d.detecte);
    const pdf = d.pdf ? ` &middot; <a class="pdf" href="${echapper(d.pdf)}" target="_blank" rel="noopener">PDF enregistré</a>` : "";
    return `<article class="fiche">
      <div class="src">${echapper(d.source)}</div>
      <h2>${titre}</h2>
      ${res}
      <div class="bas"><span>${dateArt}</span>${pdf}</div>
    </article>`;
  }).join("");
}

recherche.addEventListener("input", rendre);
rendre();
</script>
</body>
</html>"""

if __name__ == "__main__":
    main()
